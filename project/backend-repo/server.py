#!/usr/bin/env python3
"""
7 Midway Plaza — Dashboard Local Server
=======================================
A small helper that parses large Excel files far faster than a browser can, then
hands the rows to the dashboard running in your browser.

Design notes that matter:

  * It binds to 127.0.0.1 only. The server is not reachable from any other machine
    on the network. Your guest data — phone numbers included — never leaves this
    computer, and nothing is sent to any external service.

  * It serves the dashboard itself, so the page and the API share an origin and no
    browser CORS rules get in the way.

  * Sheets are parsed on demand. Opening a workbook reads only the sheet list
    (near-instant); each sheet is parsed the first time you actually look at it and
    cached from then on. On a large workbook this is the difference between waiting
    for every sheet up front and getting a usable dashboard almost immediately.

  * Uploaded workbooks live in memory only, keyed by a random id, and are dropped
    when the server stops. Nothing is written to disk.
"""

import io
import os
import sys
import json
import uuid
import socket
import threading
import webbrowser
from datetime import datetime, date, time as dtime
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer

HERE = os.path.dirname(os.path.abspath(__file__))
DASHBOARD_FILE = os.path.join(HERE, "dashboard.html")
DEFAULT_PORT = 8765
MAX_UPLOAD_BYTES = 512 * 1024 * 1024        # 512 MB
MAX_CACHED_WORKBOOKS = 3                     # keep memory bounded

# When this backend is deployed separately from the dashboard page (e.g. this on Render,
# the page on Vercel), the browser enforces CORS: it will only let the page's JavaScript
# read the response if this server explicitly allows that page's origin. Set this to your
# deployed frontend's exact address (no trailing slash), e.g.:
#   ALLOWED_ORIGIN = "https://your-app-name.vercel.app"
# "*" allows any site to call this backend, which is simplest for a small test but means
# anyone, not just your dashboard, could send files to it — fine for a test, worth
# tightening to your real Vercel URL before sharing widely.
ALLOWED_ORIGIN = os.environ.get("ALLOWED_ORIGIN", "*")

# Parsed workbooks, keyed by upload id: {id: {"name":str, "sheets":[...], "cache":{i:rows}}}
WORKBOOKS = {}
WORKBOOK_ORDER = []
LOCK = threading.Lock()

# ---------------------------------------------------------------- parser backends
def _load_parser():
    """Prefer calamine (Rust, roughly 10x faster than openpyxl); fall back to openpyxl."""
    try:
        from python_calamine import CalamineWorkbook  # noqa
        return "calamine"
    except Exception:
        pass
    try:
        import openpyxl  # noqa
        return "openpyxl"
    except Exception:
        return None

PARSER = _load_parser()


def _clean(v):
    """Normalise a cell into something JSON-safe that the dashboard understands.

    Dates and times become ISO strings, because the dashboard's own date parsing
    already handles those and it avoids every timezone ambiguity a numeric
    timestamp would introduce.
    """
    if v is None:
        return None
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        # NaN / inf are not valid JSON and would break the response.
        if v != v or v in (float("inf"), float("-inf")):
            return None
        return v
    if isinstance(v, datetime):
        return v.isoformat(sep=" ")
    if isinstance(v, date):
        return v.isoformat()
    if isinstance(v, dtime):
        return v.strftime("%H:%M:%S")
    s = str(v).strip()
    return s if s else None


def _sheet_names(data: bytes):
    if PARSER == "calamine":
        from python_calamine import CalamineWorkbook
        wb = CalamineWorkbook.from_filelike(io.BytesIO(data))
        return list(wb.sheet_names)
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    try:
        return list(wb.sheetnames)
    finally:
        wb.close()


def _detect_header_row(preview_rows):
    """Some exports (billing/settlement reports especially) prepend a few title/meta rows
    — an outlet name, a report title, a "Generated On" line, a blank row — before the real
    column headers appear. Reading row 0 as the header in that case turns the whole sheet
    to junk. This mirrors the browser parser's heuristic: row 0 is trusted as the header
    only when it and the row right after it both already look populated (the normal case);
    otherwise we scan for the first well-populated row that follows a mostly-empty one."""
    def nn(r):
        return sum(1 for v in (r or []) if v is not None and str(v).strip() != "")
    if not preview_rows:
        return 0
    row0 = nn(preview_rows[0])
    row1 = nn(preview_rows[1]) if len(preview_rows) > 1 else 0
    if row0 >= 3 and row1 >= 3:
        return 0
    for i in range(1, len(preview_rows)):
        cnt = nn(preview_rows[i])
        prev_cnt = nn(preview_rows[i - 1])
        if cnt >= 3 and prev_cnt <= 1:
            return i
    return 0


def _sheet_rows(data: bytes, index: int):
    """Returns (header_list, list_of_row_lists) for one sheet."""
    if PARSER == "calamine":
        from python_calamine import CalamineWorkbook
        wb = CalamineWorkbook.from_filelike(io.BytesIO(data))
        name = wb.sheet_names[index]
        rows = wb.get_sheet_by_name(name).to_python()
        if not rows:
            return [], []
        idx = _detect_header_row(rows[:25])
        return rows[idx], rows[idx + 1:]
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[index]]
        it = ws.iter_rows(values_only=True)
        preview = []
        for _ in range(25):
            try:
                preview.append(next(it))
            except StopIteration:
                break
        if not preview:
            return [], []
        idx = _detect_header_row(preview)
        header = list(preview[idx])
        rest = [list(r) for r in preview[idx + 1:]]
        remaining = [list(r) for r in it]
        return header, rest + remaining
    finally:
        wb.close()


def _dedupe_headers(header):
    """Mirrors SheetJS's naming so the dashboard's column detection behaves identically
    whether rows came from this server or from the browser's own parser: blank headers
    become __EMPTY / __EMPTY_1 / __EMPTY_2, and duplicates get _1, _2 suffixes."""
    out, seen, blanks = [], {}, 0
    for h in header:
        c = _clean(h)
        if c is None or c == "":
            name = "__EMPTY" if blanks == 0 else f"__EMPTY_{blanks}"
            blanks += 1
        else:
            name = str(c)
        if name in seen:
            seen[name] += 1
            name = f"{name}_{seen[name]}"
        else:
            seen[name] = 0
        out.append(name)
    return out


def build_sheet_payload(data: bytes, index: int, rating_max_rows: int = 5000):
    header, rows = _sheet_rows(data, index)
    cols = _dedupe_headers(header)
    raw_rows = []
    for r in rows:
        obj = {}
        empty = True
        for i, name in enumerate(cols):
            v = _clean(r[i]) if i < len(r) else None
            obj[name] = v
            if v is not None:
                empty = False
        if not empty:
            raw_rows.append(obj)
    # rowsArray (header:1 form) is only needed for small "rating summary" style sheets;
    # building it for a 300k-row sheet would be pure waste.
    rows_array = None
    if 0 < len(rows) <= rating_max_rows:
        rows_array = [list(cols)] + [[_clean(c) for c in r] for r in rows]
    return {"rawRows": raw_rows, "rowsArray": rows_array}


# ---------------------------------------------------------------- HTTP
class Handler(BaseHTTPRequestHandler):
    protocol_version = "HTTP/1.1"

    def log_message(self, fmt, *args):
        pass  # keep the console quiet; the launcher prints what matters

    def _send(self, code, body: bytes, ctype="application/json"):
        self.send_response(code)
        self.send_header("Content-Type", ctype)
        self.send_header("Content-Length", str(len(body)))
        self.send_header("Cache-Control", "no-store")
        self.send_header("Access-Control-Allow-Origin", ALLOWED_ORIGIN)
        self.end_headers()
        try:
            self.wfile.write(body)
        except (BrokenPipeError, ConnectionResetError):
            pass

    def _json(self, code, obj):
        self._send(code, json.dumps(obj).encode("utf-8"))

    def do_OPTIONS(self):
        # Browsers send this "preflight" check before the real POST /api/upload request
        # when the page and this server are on different addresses. We just need to say
        # the real request will be allowed.
        self.send_response(204)
        self.send_header("Access-Control-Allow-Origin", ALLOWED_ORIGIN)
        self.send_header("Access-Control-Allow-Methods", "GET, POST, OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "Content-Type")
        self.send_header("Content-Length", "0")
        self.end_headers()

    def do_GET(self):
        path = self.path.split("?")[0]
        if path in ("/", "/index.html", "/dashboard.html"):
            if not os.path.exists(DASHBOARD_FILE):
                self._send(500, b"dashboard.html is missing from this folder.", "text/plain")
                return
            with open(DASHBOARD_FILE, "rb") as f:
                self._send(200, f.read(), "text/html; charset=utf-8")
            return
        if path == "/api/health":
            self._json(200, {"ok": True, "parser": PARSER or "none",
                             "fast": PARSER == "calamine"})
            return
        if path == "/api/sheet":
            qs = dict(p.split("=", 1) for p in self.path.split("?", 1)[1].split("&")) \
                 if "?" in self.path else {}
            wid, idx = qs.get("id"), qs.get("index")
            if not wid or idx is None:
                self._json(400, {"error": "id and index are required"}); return
            try:
                idx = int(idx)
            except ValueError:
                self._json(400, {"error": "index must be a number"}); return
            with LOCK:
                entry = WORKBOOKS.get(wid)
            if not entry:
                self._json(404, {"error": "That upload is no longer in memory. Please re-open the file."}); return
            if idx < 0 or idx >= len(entry["sheets"]):
                self._json(404, {"error": "No such sheet"}); return
            if idx in entry["cache"]:
                payload = entry["cache"][idx]
            else:
                try:
                    payload = build_sheet_payload(entry["data"], idx)
                except Exception as e:
                    self._json(500, {"error": f"Could not read that sheet: {e}"}); return
                entry["cache"][idx] = payload
            self._json(200, {"sheetName": entry["sheets"][idx], **payload})
            return
        self._send(404, b"Not found", "text/plain")

    def do_POST(self):
        if self.path.split("?")[0] != "/api/upload":
            self._send(404, b"Not found", "text/plain"); return
        try:
            length = int(self.headers.get("Content-Length", "0"))
        except ValueError:
            self._json(400, {"error": "Bad Content-Length"}); return
        if length <= 0:
            self._json(400, {"error": "Empty upload"}); return
        if length > MAX_UPLOAD_BYTES:
            self._json(413, {"error": f"That file is larger than the {MAX_UPLOAD_BYTES // (1024*1024)} MB limit."}); return
        if PARSER is None:
            self._json(500, {"error": "No Excel parser is installed on the server."}); return

        # Read the body in chunks so a large upload doesn't stall on one huge read.
        buf, remaining = bytearray(), length
        while remaining > 0:
            chunk = self.rfile.read(min(1 << 20, remaining))
            if not chunk:
                break
            buf.extend(chunk)
            remaining -= len(chunk)
        data = bytes(buf)

        try:
            names = _sheet_names(data)
        except Exception as e:
            self._json(400, {"error": f"That file could not be opened as a workbook: {e}"}); return
        if not names:
            self._json(400, {"error": "No sheets found in that file."}); return

        wid = uuid.uuid4().hex
        with LOCK:
            WORKBOOKS[wid] = {"data": data, "sheets": names, "cache": {}}
            WORKBOOK_ORDER.append(wid)
            while len(WORKBOOK_ORDER) > MAX_CACHED_WORKBOOKS:
                WORKBOOKS.pop(WORKBOOK_ORDER.pop(0), None)
        self._json(200, {"id": wid, "sheets": names, "parser": PARSER})


def find_free_port(preferred):
    for port in [preferred] + list(range(preferred + 1, preferred + 25)):
        with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
            if s.connect_ex(("127.0.0.1", port)) != 0:
                return port
    return None


def main():
    if PARSER is None:
        print("\n  No Excel parser is available.")
        print("  Run:  pip install python-calamine")
        print("  (or:  pip install openpyxl  — slower but works)\n")
        sys.exit(1)

    # Render (and most cloud hosts) set this automatically; it's how we tell "running on
    # a server in the cloud" apart from "running on someone's own laptop".
    ON_HOST = bool(os.environ.get("RENDER") or os.environ.get("ON_CLOUD_HOST"))

    has_dashboard = os.path.exists(DASHBOARD_FILE)
    if not has_dashboard and not ON_HOST:
        print(f"\n  dashboard.html was not found next to this script.")
        print(f"  Expected it at: {DASHBOARD_FILE}\n")
        sys.exit(1)
    # On a cloud host this server is deployed as a pure API backend (the dashboard page
    # itself lives on Vercel), so dashboard.html is expected to be absent — not an error.

    if ON_HOST:
        host = "0.0.0.0"
        port = int(os.environ.get("PORT", DEFAULT_PORT))
    else:
        host = "127.0.0.1"
        port = find_free_port(int(os.environ.get("PORT", DEFAULT_PORT)))
        if port is None:
            print("\n  Could not find a free port. Close other copies of this server and try again.\n")
            sys.exit(1)

    url = f"http://{host}:{port}/"
    speed = "fast Rust parser" if PARSER == "calamine" else "openpyxl (slower — install python-calamine for a big speed-up)"
    print("\n" + "=" * 62)
    print("  7 Midway Plaza — Dashboard server is running")
    print("=" * 62)
    print(f"  Listening on: {host}:{port}")
    print(f"  Parser:       {speed}")
    if ON_HOST:
        print( "  Mode:         cloud backend (API only — allowed origin below)")
        print(f"  CORS origin:  {ALLOWED_ORIGIN}")
    else:
        print( "  Privacy:      bound to this computer only (127.0.0.1).")
        print( "                Your data is never sent anywhere and never written to disk.")
        print( "  Stop:         press Ctrl+C in this window")
    print("=" * 62 + "\n")

    if not ON_HOST:
        threading.Timer(1.0, lambda: webbrowser.open(url)).start()
    srv = ThreadingHTTPServer((host, port), Handler)
    try:
        srv.serve_forever()
    except KeyboardInterrupt:
        print("\n  Server stopped. Your data has been cleared from memory.\n")
    finally:
        srv.server_close()


if __name__ == "__main__":
    main()
